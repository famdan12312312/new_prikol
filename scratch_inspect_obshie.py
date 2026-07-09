import pandas as pd
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

f_name = 'individ_plan_Архипов.xlsx'
wb = openpyxl.load_workbook(f_name, data_only=True)
sheet = wb['Общие сведения']

print("=== Raw Cells in 'Общие сведения' ===")
for r in range(1, 25):
    row_vals = [f"{sheet.cell(row=r, column=c).coordinate}: '{sheet.cell(row=r, column=c).value}'" for c in range(1, 10) if sheet.cell(row=r, column=c).value is not None]
    if row_vals:
        print(f"Row {r}:", ", ".join(row_vals))

print("\n=== Transposed DataFrame Values ===")
df = pd.read_excel(f_name, sheet_name='Общие сведения', header=None)
df_t = df.T

for r in range(min(15, df_t.shape[0])):
    row_vals = [f"Col {c}: '{df_t.iloc[r, c]}'" for c in range(df_t.shape[1]) if pd.notna(df_t.iloc[r, c]) and str(df_t.iloc[r, c]).strip()]
    if row_vals:
        print(f"Transposed Row {r}:", ", ".join(row_vals))
