import openpyxl
import re

wb = openpyxl.load_workbook('individ_plan_Архипов.xlsx', data_only=True)
pattern = re.compile(r'\d{2}\.\d{2}\.\d{2}')

found = []
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None and isinstance(val, str):
                match = pattern.search(val)
                if match:
                    found.append((sheet_name, cell := sheet.cell(row=r, column=c).coordinate, val))

with open("scratch_direction_found.txt", "w", encoding="utf-8") as f_out:
    for f in found:
        f_out.write(f"Sheet: {f[0]}, Cell: {f[1]}, Value: {f[2]}\n")
print(f"Done, found matches: {len(found)}")
