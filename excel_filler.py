import os
import re
import openpyxl
from openpyxl.styles import Font, Border, PatternFill, Alignment, Side

def copy_cell_style(src_cell, dest_cell):
    """
    Копирует стиль из src_cell в dest_cell.
    """
    if src_cell.has_style:
        dest_cell.font = Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            color=src_cell.font.color,
            underline=src_cell.font.underline,
            strike=src_cell.font.strike
        )
        dest_cell.fill = PatternFill(
            fill_type=src_cell.fill.fill_type,
            start_color=src_cell.fill.start_color,
            end_color=src_cell.fill.end_color
        )
        dest_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        )
        dest_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            text_rotation=src_cell.alignment.text_rotation,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        )
        dest_cell.number_format = src_cell.number_format

def read_cell_value(ws, row, col):
    """
    Безопасно считывает значение ячейки ws.cell(row, col).
    Если ячейка является MergedCell, возвращает значение её левой верхней ячейки.
    """
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

def write_cell(ws, row, col, val):
    """
    Безопасно записывает значение в ячейку ws.cell(row, col).
    Если ячейка защищена от записи (например, объединенная MergedCell),
    находит её левую верхнюю ячейку и записывает значение туда.
    """
    cell = ws.cell(row=row, column=col)
    try:
        cell.value = val
    except (AttributeError, TypeError):
        # Ищем начало объединенной ячейки
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                try:
                    ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = val
                    return
                except:
                    pass

def get_semester_season(sem_val):
    """
    Определяет сезон семестра: 'autumn' (нечетный/осенний) или 'spring' (четный/весенний).
    """
    sem_str = str(sem_val).strip().lower()
    if not sem_str:
        return 'autumn'
        
    # Ищем все цифры в строке семестра
    digits = re.findall(r'\d+', sem_str)
    if digits:
        try:
            num = int(digits[0])
            if num % 2 != 0:
                return 'autumn'
            else:
                return 'spring'
        except ValueError:
            pass
            
    if "осен" in sem_str or "нечет" in sem_str:
        return 'autumn'
    elif "весен" in sem_str or "чет" in sem_str:
        return 'spring'
        
    return 'autumn'

def map_columns(ws, header_row):
    """
    Сканирует область заголовка и сопоставляет ключевые слова с номерами колонок.
    Для надежности сканируются все строки от 1 до header_row + 2, чтобы корректно собрать 
    текст из длинных вертикально объединенных ячеек.
    """
    col_map = {}
    for c in range(1, ws.max_column + 1):
        parts = []
        for r in range(1, header_row + 3):
            cell_val = read_cell_value(ws, r, c)
            if cell_val:
                parts.append(str(cell_val).strip().lower())
        val_str = " ".join(parts)
        
        # 1. Дисциплина / Наименование
        if "дисциплин" in val_str or "наимен" in val_str or "предмет" in val_str or "вид работы" in val_str:
            col_map["subject"] = c
            
        # 2. Группа
        elif "групп" in val_str:
            col_map["group"] = c
            
        # 3. Направление / Специальность
        elif "направл" in val_str or "спец" in val_str:
            col_map["direction"] = c
            
        # 4. Семестр
        elif "семестр" in val_str or re.search(r'\bсем-р\b', val_str):
            col_map["semester"] = c
            
        # 5. Лекции
        elif "лекц" in val_str or "лек" in val_str or val_str.strip() == "л":
            col_map["lectures"] = c
            
        # 6. Лабораторные
        elif "лаб" in val_str or "л.р." in val_str or "л/р" in val_str or re.search(r'\bлр\b', val_str):
            col_map["laboratories"] = c
            
        # 7. Консультации
        elif "консульт" in val_str or "конс" in val_str:
            col_map["consultations"] = c
            
        # 8. Экзамены
        elif "экзам" in val_str or "экз" in val_str:
            col_map["exams"] = c
            
        # 9. Зачеты
        elif "зачет" in val_str or "зач" in val_str:
            col_map["zachets"] = c
            
        # 10. Курсовые работы (КП / КР)
        elif "курсов" in val_str or re.search(r'\b(кп|кр)\b', val_str):
            col_map["coursework"] = c
            
        # 11. ВКР / Диплом
        elif "вкр" in val_str or "диплом" in val_str or "выпускн" in val_str:
            col_map["vkr"] = c
            
        # 12. ГЭК
        elif "гэк" in val_str or "гэка" in val_str or "гос" in val_str:
            col_map["gek"] = c
            
        # 13. Дополнительные часы / Доп
        elif "доп" in val_str or "сверх" in val_str or "дополн" in val_str or "дополнит" in val_str:
            col_map["additional"] = c
            
        # 14. Всего / Итого
        elif "всего" in val_str or "итого" in val_str:
            col_map["total"] = c
            
        # 15. Практические / Семинары / Руководство практикой (Разделение)
        # Сначала проверяем на практику (учебная, производственная, руководство практикой)
        elif any(kw in val_str for kw in ["руковод", "произв", "учебн", "производств", "рук. практ", "рук-во", "преддиплом", "производственная"]):
            col_map["practice"] = c
        elif val_str.strip() == "практика":
            col_map["practice"] = c
        # Затем на практические занятия (семинары, упражнения)
        elif any(kw in val_str for kw in ["семин", "сем", "заняти", "практ"]):
            col_map["practicals"] = c
        elif re.search(r'\bпр\b|\bпр\.', val_str):
            col_map["practicals"] = c
    return col_map

def fill_fixed_rows(ws, row_indices, loads_to_write, col_map):
    """
    Заполняет отведенные строки шаблона без изменения структуры (без вставки новых строк).
    Очищает оставшиеся пустые строки в выделенной зоне.
    """
    L = len(loads_to_write)
    for i, r in enumerate(row_indices):
        if i < L:
            load_item = loads_to_write[i]
            
            write_cell(ws, r, col_map["subject"], load_item.get("subject", ""))
            if "group" in col_map:
                write_cell(ws, r, col_map["group"], load_item.get("group", ""))
            if "direction" in col_map:
                write_cell(ws, r, col_map["direction"], load_item.get("direction", ""))
            if "semester" in col_map:
                write_cell(ws, r, col_map["semester"], load_item.get("semester", ""))
                
            hours = load_item.get("hours", {})
            hours_keys_map = {
                "lectures": "lectures",
                "practicals": "practicals",
                "laboratories": "laboratories",
                "consultations": "consultations",
                "exams": "exams",
                "zachets": "zachets",
                "coursework": "coursework",
                "practice": "practice",
                "vkr": "vkr",
                "gek": "gek",
                "additional": "additional"
            }
            
            for db_key, col_key in hours_keys_map.items():
                if col_key in col_map:
                    h_val = float(hours.get(db_key, 0.0))
                    write_cell(ws, r, col_map[col_key], h_val if h_val > 0.0 else None)
                    
            if "total" in col_map:
                tot_val = float(load_item.get("total", 0.0))
                write_cell(ws, r, col_map["total"], tot_val if tot_val > 0.0 else None)
        else:
            # Очищаем неиспользованные шаблонные строки в этой секции, сохраняя форматирование
            write_cell(ws, r, col_map["subject"], None)
            if "group" in col_map:
                write_cell(ws, r, col_map["group"], None)
            if "direction" in col_map:
                write_cell(ws, r, col_map["direction"], None)
            if "semester" in col_map:
                write_cell(ws, r, col_map["semester"], None)
            
            for key in ["lectures", "practicals", "laboratories", "consultations", "exams", "zachets", "coursework", "practice", "vkr", "gek", "additional", "total"]:
                if key in col_map:
                    write_cell(ws, r, col_map[key], None)

def update_total_row_sums(ws, row_indices, total_row, col_map):
    """
    Подсчитывает суммы часов по указанным строкам и записывает в строку итогов (если там нет формул).
    """
    if total_row and total_row <= ws.max_row:
        for key in ["lectures", "practicals", "laboratories", "consultations", "exams", "zachets", "coursework", "practice", "vkr", "gek", "additional", "total"]:
            if key in col_map:
                col_idx = col_map[key]
                target_val = read_cell_value(ws, total_row, col_idx)
                
                if target_val is None or not str(target_val).startswith("="):
                    col_sum = 0.0
                    for r in row_indices:
                        val = read_cell_value(ws, r, col_idx)
                        try:
                            col_sum += float(val) if val is not None else 0.0
                        except:
                            pass
                    write_cell(ws, total_row, col_idx, col_sum if col_sum > 0.0 else None)

STRICT_PLAN_COORDINATES = {
    "employee_fio": {"coordinate": "D3", "sheet": "Общие сведения"},
    "employee_fio_title": {"coordinate": "A31", "sheet": "Тит. лист"},
    "employee_position": {"coordinate": "D4", "sheet": "Общие сведения"},
    "employee_position_title": {"coordinate": "A34", "sheet": "Тит. лист"},
    "employee_rate": {"coordinate": "D5", "sheet": "Общие сведения"},
    "employee_conditions": {"coordinate": "D6", "sheet": "Общие сведения"},
    "employee_degree": {"coordinate": "D7", "sheet": "Общие сведения"},
    "employee_title": {"coordinate": "D8", "sheet": "Общие сведения"},
    "employee_contract": {"coordinate": "D10", "sheet": "Общие сведения"},
    "employee_contract_duration": {"coordinate": "D11", "sheet": "Общие сведения"},
    "employee_edu_inst_year": {"coordinate": "D12", "sheet": "Общие сведения"},
    "employee_edu_specialty_1": {"coordinate": "D14", "sheet": "Общие сведения"},
    "employee_edu_specialty_2": {"coordinate": "D15", "sheet": "Общие сведения"},
    "employee_edu_qualification": {"coordinate": "D18", "sheet": "Общие сведения"},
    "department_name": {"coordinate": "A25", "sheet": "Тит. лист"},
    "department_head": {"coordinate": "F16", "sheet": "Тит. лист"},
    "institute_name": {"coordinate": "A28", "sheet": "Тит. лист"},
    "study_year": {"coordinate": "A37", "sheet": "Тит. лист"}
}

def fill_teacher_plan(template_path, output_path, new_fio=None, loads=None, search_name="Обухов", position=None, employment_conditions=None, degree=None, title=None, placeholder_map=None):
    """
    Основная функция заполнения Excel-шаблона индивидуального плана.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    
    # 1. Замена метаданных и ФИО преподавателя
    if placeholder_map:
        # Двухэтапная запись
        # Этап 1: Прямая запись по координатам
        from openpyxl.utils import coordinate_to_tuple
        for key, item in placeholder_map.items():
            val = item.get("value")
            if val is not None:
                coord_info = STRICT_PLAN_COORDINATES.get(key)
                if coord_info:
                    sheet_name = coord_info["sheet"]
                    coord = coord_info["coordinate"]
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        try:
                            row_idx, col_idx = coordinate_to_tuple(coord)
                            write_cell(ws, row_idx, col_idx, val)
                        except Exception:
                            pass
                            
        # Этап 2: Глобальный поиск и замена заглушек
        for key, item in placeholder_map.items():
            placeholder = str(item.get("placeholder") or "").strip()
            val = str(item.get("value") or "").strip()
            if placeholder and val:
                pattern = re.compile(re.escape(placeholder), re.IGNORECASE)
                for s_name in wb.sheetnames:
                    ws = wb[s_name]
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            cell_val = read_cell_value(ws, r, c)
                            if cell_val and isinstance(cell_val, str):
                                if placeholder.lower() in cell_val.lower():
                                    new_val = pattern.sub(val, cell_val)
                                    write_cell(ws, r, c, new_val)
    else:
        # Старая обратная совместимость
        pattern = re.compile(re.escape(search_name) + r"(\s+[А-Яа-я]\.\s*[А-Яа-я]\.)?", re.IGNORECASE)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = read_cell_value(ws, r, c)
                    if val and isinstance(val, str):
                        if search_name.lower() in val.lower():
                            new_val = pattern.sub(new_fio, val)
                            write_cell(ws, r, c, new_val)
                            
            # Запись метаданных преподавателя
            for r in range(1, min(16, ws.max_row + 1)):
                c0_val = str(read_cell_value(ws, r, 1) or "").strip().lower()
                if "должность" in c0_val and position:
                    write_cell(ws, r, 5, position)
                elif "условия привлечения" in c0_val and employment_conditions:
                    write_cell(ws, r, 5, employment_conditions)
                elif "ученая степень" in c0_val and degree:
                    write_cell(ws, r, 5, degree)
                elif "ученое звание" in c0_val and title:
                    write_cell(ws, r, 5, title)
                    
    
# 2. Заполнение таблиц нагрузки
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Поиск строки заголовка
        header_row = None
        for r in range(1, min(40, ws.max_row + 1)):
            row_vals = [str(read_cell_value(ws, r, c) or "").strip().lower() for c in range(1, ws.max_column + 1)]
            has_sub = any("дисциплин" in v or "наимен" in v or "предмет" in v or "вид работы" in v for v in row_vals)
            has_grp = any("групп" in v for v in row_vals)
            if has_sub and has_grp:
                header_row = r
                break
                
        if not header_row:
            continue
            
        col_map = map_columns(ws, header_row)
        if "subject" not in col_map:
            continue
            
        # Определяем семестровую специфику листа (из имени)
        sheet_season = None
        if any(kw in sheet_name.lower() for kw in ["1", "осен", "autumn", "нечет"]):
            sheet_season = "autumn"
        elif any(kw in sheet_name.lower() for kw in ["2", "весен", "spring", "чет"]):
            sheet_season = "spring"
            
        # Разделяем учебную нагрузку преподавателя по сезонам
        autumn_loads = [l for l in loads if get_semester_season(l.get("semester", "1")) == "autumn"]
        spring_loads = [l for l in loads if get_semester_season(l.get("semester", "1")) == "spring"]
        
        # Проверяем, есть ли на листе весенняя секция вообще
        has_spring_section = False
        for r in range(header_row + 1, ws.max_row + 1):
            row_vals = [str(read_cell_value(ws, r, c) or "").strip().lower() for c in range(1, ws.max_column + 1)]
            row_str = " ".join(row_vals)
            if "весенний семестр" in row_str or "б) весен" in row_str:
                has_spring_section = True
                break

        # Сканируем строки для выделения пустых зон осенней и весенней таблиц
        autumn_rows = []
        spring_rows = []
        zone = "autumn"
        
        for r in range(header_row + 1, ws.max_row + 1):
            row_vals = [str(read_cell_value(ws, r, c) or "").strip().lower() for c in range(1, ws.max_column + 1)]
            row_str = " ".join(row_vals)
            
            # Проверяем наличие разделителя весеннего семестра
            if "весенний семестр" in row_str or "б) весен" in row_str:
                zone = "spring"
                continue
                
            # Если встретили строку итогов/всего
            if any(hw in row_str for hw in ["итого", "всего", "выполнено"]):
                if zone == "spring" and spring_rows:
                    break
                elif zone == "autumn" and autumn_rows and not has_spring_section:
                    break
                continue
                
            # Пропускаем служебные заголовки
            if any(hw in row_str for hw in ["дисциплина", "лекции", "семестр", "группа", "весенний", "осенний"]):
                continue
                
            if zone == "autumn":
                autumn_rows.append(r)
            else:
                spring_rows.append(r)
                
        # Находим строки итогов для каждой секции
        autumn_total_row = None
        if autumn_rows:
            for r in range(autumn_rows[-1] + 1, ws.max_row + 1):
                row_vals = [str(read_cell_value(ws, r, c) or "").strip().lower() for c in range(1, ws.max_column + 1)]
                if any("итого" in v or "всего" in v for v in row_vals):
                    autumn_total_row = r
                    break
                    
        spring_total_row = None
        if spring_rows:
            for r in range(spring_rows[-1] + 1, ws.max_row + 1):
                row_vals = [str(read_cell_value(ws, r, c) or "").strip().lower() for c in range(1, ws.max_column + 1)]
                if any("итого" in v or "всего" in v for v in row_vals):
                    spring_total_row = r
                    break
                    
        # Заполнение таблиц нагрузки в выделенные пустые зоны
        if spring_rows:
            # На листе есть две секции (осенняя и весенняя вертикально)
            fill_fixed_rows(ws, autumn_rows, autumn_loads, col_map)
            update_total_row_sums(ws, autumn_rows, autumn_total_row, col_map)
            
            fill_fixed_rows(ws, spring_rows, spring_loads, col_map)
            update_total_row_sums(ws, spring_rows, spring_total_row, col_map)
        else:
            # На листе одна общая таблица (например, на раздельных листах семестров)
            if sheet_season == "spring":
                fill_fixed_rows(ws, autumn_rows, spring_loads, col_map)
            elif not sheet_season:
                fill_fixed_rows(ws, autumn_rows, loads, col_map)
            else:
                fill_fixed_rows(ws, autumn_rows, autumn_loads, col_map)
                
            update_total_row_sums(ws, autumn_rows, autumn_total_row, col_map)

        # 3. Заполнение перечня преподаваемых дисциплин на листе "Общие сведения"
    if "Общие сведения" in wb.sheetnames and loads:
        ws_info = wb["Общие сведения"]
        unique_subjects = []
        for l in loads:
            subj = str(l.get("subject", "")).strip()
            if subj and subj not in unique_subjects:
                if not any(kw in subj.lower() for kw in ["итого", "всего", "фактически выполнено"]):
                    unique_subjects.append(subj)
        
        start_row = None
        for r in range(1, ws_info.max_row + 1):
            val = str(read_cell_value(ws_info, r, 1) or "").strip().lower()
            if "перечень преподаваемых дисциплин" in val:
                start_row = r
                break
                
        if start_row:
            for r in range(start_row + 1, start_row + 30):
                write_cell(ws_info, r, 1, None)
            for idx, subj in enumerate(unique_subjects):
                write_cell(ws_info, start_row + 1 + idx, 1, f"{idx + 1}. {subj}")
                

    wb.save(output_path)
    return output_path
