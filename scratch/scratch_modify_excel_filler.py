filepath = "excel_filler.py"
with open(filepath, "r", encoding="utf-8") as f:
    content = f.read()

start_marker = "def fill_teacher_plan("
end_marker = "# 2. Заполнение таблиц нагрузки"

start_idx = content.find(start_marker)
if start_idx == -1:
    print("Error: start_marker not found")
    exit(1)

end_idx = content.find(end_marker)
if end_idx == -1:
    print("Error: end_marker not found")
    exit(1)

STRICT_PLAN_COORDINATES_CODE = """STRICT_PLAN_COORDINATES = {
    "employee_fio": {"coordinate": "B2", "sheet": "Общие сведения"},
    "employee_position": {"coordinate": "C3", "sheet": "Общие сведения"},
    "employee_rate": {"coordinate": "D5", "sheet": "Общие сведения"},
    "employee_conditions": {"coordinate": "D6", "sheet": "Общие сведения"},
    "employee_degree": {"coordinate": "D7", "sheet": "Общие сведения"},
    "employee_title": {"coordinate": "D8", "sheet": "Общие сведения"},
    "employee_contract_num": {"coordinate": "E9", "sheet": "Общие сведения"},
    "employee_contract_date": {"coordinate": "F9", "sheet": "Общие сведения"},
    "employee_contract_duration": {"coordinate": "G10", "sheet": "Общие сведения"},
    "employee_edu_institution": {"coordinate": "E11", "sheet": "Общие сведения"},
    "employee_edu_year": {"coordinate": "F12", "sheet": "Общие сведения"},
    "employee_edu_specialty": {"coordinate": "G15", "sheet": "Общие сведения"},
    "employee_edu_qualification": {"coordinate": "H18", "sheet": "Общие сведения"},
    "department_name": {"coordinate": "A25", "sheet": "Тит. лист"},
    "department_head": {"coordinate": "F16", "sheet": "Тит. лист"},
    "institute_name": {"coordinate": "A28", "sheet": "Тит. лист"}
}
"""

new_function_code = """def fill_teacher_plan(template_path, output_path, new_fio=None, loads=None, search_name="Обухов", position=None, employment_conditions=None, degree=None, title=None, placeholder_map=None):
    \"\"\"
    Основная функция заполнения Excel-шаблона индивидуального плана.
    \"\"\"
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    
    # 1. Замена метаданных и ФИО преподавателя
    if placeholder_map:
        # Двухэтапная запись
        # Этап 1: Прямая запись по координатам
        from openpyxl.utils import coordinate_to_tuple
        for key, item in placeholder_map.items():
            val = item.get("value")
            if val is not None:
                coord_info = STRICT_PLAN_COORDINATES.get(key)
                if coord_info:
                    sheet_name = coord_info["sheet"]
                    coord = coord_info["coordinate"]
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        try:
                            row_idx, col_idx = coordinate_to_tuple(coord)
                            write_cell(ws, row_idx, col_idx, val)
                        except Exception:
                            pass
                            
        # Этап 2: Глобальный поиск и замена заглушек
        for key, item in placeholder_map.items():
            placeholder = str(item.get("placeholder") or "").strip()
            val = str(item.get("value") or "").strip()
            if placeholder and val:
                pattern = re.compile(re.escape(placeholder), re.IGNORECASE)
                for s_name in wb.sheetnames:
                    ws = wb[s_name]
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            cell_val = read_cell_value(ws, r, c)
                            if cell_val and isinstance(cell_val, str):
                                if placeholder.lower() in cell_val.lower():
                                    new_val = pattern.sub(val, cell_val)
                                    write_cell(ws, r, c, new_val)
    else:
        # Старая обратная совместимость
        pattern = re.compile(re.escape(search_name) + r"(\\s+[А-Яа-я]\\.\\s*[А-Яа-я]\\.)?", re.IGNORECASE)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = read_cell_value(ws, r, c)
                    if val and isinstance(val, str):
                        if search_name.lower() in val.lower():
                            new_val = pattern.sub(new_fio, val)
                            write_cell(ws, r, c, new_val)
                            
            # Запись метаданных преподавателя
            for r in range(1, min(16, ws.max_row + 1)):
                c0_val = str(read_cell_value(ws, r, 1) or "").strip().lower()
                if "должность" in c0_val and position:
                    write_cell(ws, r, 5, position)
                elif "условия привлечения" in c0_val and employment_conditions:
                    write_cell(ws, r, 5, employment_conditions)
                elif "ученая степень" in c0_val and degree:
                    write_cell(ws, r, 5, degree)
                elif "ученое звание" in c0_val and title:
                    write_cell(ws, r, 5, title)
                    
    """

new_content = content[:start_idx] + STRICT_PLAN_COORDINATES_CODE + "\n" + new_function_code + "\n" + content[end_idx:]

# Теперь добавим запись списка предметов перед wb.save(output_path)
save_marker = "wb.save(output_path)"
save_idx = new_content.rfind(save_marker)
if save_idx == -1:
    print("Error: save_marker not found")
    exit(1)

disciplines_list_code = """    # 3. Заполнение перечня преподаваемых дисциплин на листе "Общие сведения"
    if "Общие сведения" in wb.sheetnames and loads:
        ws_info = wb["Общие сведения"]
        unique_subjects = []
        for l in loads:
            subj = str(l.get("subject", "")).strip()
            if subj and subj not in unique_subjects:
                if not any(kw in subj.lower() for kw in ["итого", "всего", "фактически выполнено"]):
                    unique_subjects.append(subj)
        
        start_row = None
        for r in range(1, ws_info.max_row + 1):
            val = str(read_cell_value(ws_info, r, 1) or "").strip().lower()
            if "перечень преподаваемых дисциплин" in val:
                start_row = r
                break
                
        if start_row:
            for r in range(start_row + 1, start_row + 30):
                write_cell(ws_info, r, 1, None)
            for idx, subj in enumerate(unique_subjects):
                write_cell(ws_info, start_row + 1 + idx, 1, f"{idx + 1}. {subj}")
                
"""

final_content = new_content[:save_idx] + disciplines_list_code + "\n    " + new_content[save_idx:]

with open(filepath, "w", encoding="utf-8") as f:
    f.write(final_content)

print("Modification of excel_filler.py successful")
