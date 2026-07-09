import openpyxl

wb = openpyxl.load_workbook('individ_plan_Архипов.xlsx', data_only=True)
sheet = wb['Тит. лист']

with open("scratch_title_cells.txt", "w", encoding="utf-8") as f_out:
    for r in range(1, sheet.max_row + 1):
        row_vals = [f"{sheet.cell(row=r, column=c).coordinate}: {sheet.cell(row=r, column=c).value}" for c in range(1, sheet.max_column + 1) if sheet.cell(row=r, column=c).value is not None]
        if row_vals:
            f_out.write(f"Row {r}: " + ", ".join(row_vals) + "\n")
print("Done writing title cells")
